"""Report file generation: CSV, PDF (reportlab), XLSX (openpyxl).

All three formats are built from the same normalized row data and the same
metadata dict, so a report always shows identical figures regardless of
which format the user picked — only the presentation differs.
"""

from __future__ import annotations

import csv
import io

from django.utils import timezone

REPORT_META = {
    "inventory": {
        "title": "Inventory Summary",
        "description": "Every reagent with location, quantity and status.",
    },
    "hazard": {
        "title": "Hazard & Compliance",
        "description": "All chemicals grouped by GHS hazard classification.",
    },
    "expiry": {
        "title": "Expiry Forecast",
        "description": "Chemicals expiring within the next 90 days.",
    },
    "low_stock": {
        "title": "Low-Stock Reorder",
        "description": "Reagents at or below their minimum stock threshold.",
    },
}

HEADERS = [
    "Name", "CAS Number", "Hazard Class", "Status", "Location",
    "Quantity", "Unit", "Min Stock", "Expiry Date",
]

BRAND_COLOR = "0F766E"  # matches the app's primary/teal accent


def chemical_to_row(chemical) -> list[str]:
    return [
        chemical.name,
        chemical.cas_number or "",
        chemical.get_hazard_class_display(),
        chemical.get_status_display(),
        chemical.location or "",
        str(chemical.quantity),
        chemical.unit,
        str(chemical.min_stock),
        chemical.expiry_date.isoformat() if chemical.expiry_date else "",
    ]


def _summary_line(chemicals, generated_by) -> str:
    now = timezone.now()
    return f"Generated {now:%Y-%m-%d %H:%M} UTC by {generated_by} \u2014 {len(chemicals)} record(s)"


def build_csv(chemicals, category: str, generated_by: str) -> bytes:
    meta = REPORT_META[category]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([f"ChemSafe \u2014 {meta['title']}"])
    writer.writerow([meta["description"]])
    writer.writerow([_summary_line(chemicals, generated_by)])
    writer.writerow([])
    writer.writerow(HEADERS)
    for chemical in chemicals:
        writer.writerow(chemical_to_row(chemical))
    # utf-8-sig so Excel (which the .csv is very likely opened in) detects
    # UTF-8 correctly instead of mangling non-ASCII characters.
    return buf.getvalue().encode("utf-8-sig")


def build_pdf(chemicals, category: str, generated_by: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    meta = REPORT_META[category]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=18 * mm, bottomMargin=16 * mm, leftMargin=14 * mm, rightMargin=14 * mm,
        title=f"ChemSafe - {meta['title']}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], textColor=colors.HexColor("#0f172a"),
        fontSize=18, spaceAfter=2, alignment=0,
    )
    meta_style = ParagraphStyle(
        "ReportMeta", parent=styles["Normal"], textColor=colors.HexColor("#64748b"), fontSize=9,
    )

    elements = [
        Paragraph(f"ChemSafe \u2014 {meta['title']}", title_style),
        Paragraph(meta["description"], meta_style),
        Spacer(1, 4),
        Paragraph(_summary_line(chemicals, generated_by), meta_style),
        Spacer(1, 10),
    ]

    table_data = [HEADERS] + [chemical_to_row(c) for c in chemicals]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_COLOR}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    if not chemicals:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("No records match this report.", meta_style))

    doc.build(elements)
    return buf.getvalue()


def build_xlsx(chemicals, category: str, generated_by: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    meta = REPORT_META[category]
    wb = Workbook()
    ws = wb.active
    ws.title = meta["title"][:31]  # Excel sheet-name length limit

    last_col_letter = get_column_letter(len(HEADERS))

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = f"ChemSafe \u2014 {meta['title']}"
    ws["A1"].font = Font(size=14, bold=True, color="0F172A")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = meta["description"]
    ws["A2"].font = Font(size=10, color="64748B")

    ws.merge_cells(f"A3:{last_col_letter}3")
    ws["A3"] = _summary_line(chemicals, generated_by)
    ws["A3"].font = Font(size=9, italic=True, color="64748B")

    header_row = 5
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=BRAND_COLOR)
        cell.alignment = Alignment(horizontal="center")

    for r, chemical in enumerate(chemicals, start=header_row + 1):
        for col, value in enumerate(chemical_to_row(chemical), start=1):
            ws.cell(row=r, column=col, value=value)

    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
